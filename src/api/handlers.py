"""File-based reconciliation handler.

Receives a workbook (bytes), runs the deterministic pre-pass from
``reconcile.py``, applies the confident (zero-judgment) fixes back into the
control sheet (columns M / N), records exceptions on a dedicated sheet, and
returns the modified workbook bytes.

No Microsoft Graph API and no MCP: all I/O is local via openpyxl. The caller
(e.g. a Power Automate flow) is responsible for getting the file to this
endpoint and writing the returned bytes back to the source file.
"""

from __future__ import annotations

import io
import logging
import os
import tempfile
from typing import Any

import openpyxl

from src.reconciliation import reconcile

logger = logging.getLogger(__name__)

EXCEPTIONS_SHEET = "ReconAgent Exceptions"

EXCEPTION_COLUMNS = [
    "Control Row",
    "K3 Code",
    "Sch K-1 Line",
    "Description",
    "Current Part II",
    "Current Part X",
    "Proposed Part II",
    "Proposed Part X",
    "Why (needs judgment)",
]

# 1-based column indexes of the edit targets on the control sheet.
P2_COL = reconcile.COLS["P2"]  # column M — Sch K2 Part II line
PX_COL = reconcile.COLS["PX"]  # column N — Sch K2 Part X line


def _write_exceptions_sheet(
    workbook: openpyxl.Workbook, exceptions: list[dict[str, Any]]
) -> None:
    """Create/replace the exceptions sheet listing rows that need judgment."""
    if EXCEPTIONS_SHEET in workbook.sheetnames:
        del workbook[EXCEPTIONS_SHEET]

    sheet = workbook.create_sheet(EXCEPTIONS_SHEET)
    sheet.append(EXCEPTION_COLUMNS)

    for exc in exceptions:
        sheet.append(
            [
                exc.get("row"),
                exc.get("K3"),
                exc.get("line"),
                exc.get("desc"),
                exc.get("current_part2"),
                exc.get("current_partx"),
                exc.get("proposed_part2"),
                exc.get("proposed_partx"),
                exc.get("why"),
            ]
        )


def run_reconciliation(
    workbook_bytes: bytes, filename: str = "workbook.xlsm"
) -> dict[str, Any]:
    """Run the pre-pass on a workbook and return the modified workbook.

    Args:
        workbook_bytes: Raw bytes of the uploaded .xlsm / .xlsx workbook.
        filename: Original filename (used to preserve the macro format).

    Returns:
        {
            "workbook_bytes": bytes,   # modified workbook to write back
            "summary": {...},          # counts
            "confident": [...],        # applied changes
            "exceptions": [...],       # flagged rows
        }
    """
    is_macro = filename.lower().endswith(".xlsm")
    suffix = ".xlsm" if is_macro else ".xlsx"

    tmp_path = None
    try:
        with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
            tmp.write(workbook_bytes)
            tmp_path = tmp.name

        # 1) Deterministic pre-pass (reads control sheet via openpyxl).
        payload = reconcile.run(tmp_path)
        confident = payload["confident"]
        exceptions = payload["exceptions"]
        logger.info(
            "[run_reconciliation] %s confident, %s exceptions",
            len(confident),
            len(exceptions),
        )

        # 2) Re-open writable and apply confident fixes to columns M / N.
        workbook = openpyxl.load_workbook(tmp_path, keep_vba=is_macro)
        control = workbook[reconcile.CONTROL_SHEET]

        applied = 0
        for change in confident:
            row = change["row"]
            control.cell(row=row, column=P2_COL).value = change["proposed_part2"]
            control.cell(row=row, column=PX_COL).value = change["proposed_partx"]
            applied += 1

        # 3) Record exceptions on a dedicated sheet.
        _write_exceptions_sheet(workbook, exceptions)

        # 4) Serialize back to bytes.
        buffer = io.BytesIO()
        workbook.save(buffer)
        workbook.close()

        return {
            "workbook_bytes": buffer.getvalue(),
            "summary": {
                **payload["summary"],
                "confident_applied": applied,
            },
            "confident": confident,
            "exceptions": exceptions,
        }

    finally:
        if tmp_path and os.path.exists(tmp_path):
            try:
                os.remove(tmp_path)
            except OSError:
                logger.warning("Could not remove temp file %s", tmp_path)
